#!/usr/bin/env python3
"""
KiBot Variant Manager v4 — Windows XP style, red corporate theme
"""

import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext
import yaml, subprocess, threading, os, re, glob, json, sys, datetime, csv
from pathlib import Path
from PIL import Image, ImageTk

try:
    from tkinterdnd2 import DND_FILES, TkinterDnD
    HAS_DND = True
except ImportError:
    HAS_DND = False

try:
    from openpyxl import load_workbook
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

IS_WINDOWS = sys.platform == 'win32'
sys.setrecursionlimit(10000)
CONFIG_FILE = os.path.join(os.path.expanduser("~"), ".kibot_manager.json")
ASSETS_DIR = os.path.join(getattr(sys, '_MEIPASS', os.path.dirname(os.path.abspath(__file__))), 'assets')
POPEN_FLAGS = {}
if IS_WINDOWS:
    POPEN_FLAGS['creationflags'] = 0x08000000
KIBOT_MIN_VERSION = (1, 9, 0)
KICAD_MIN_MAJOR   = 10

# ══════════════════════════════════════════════
#  XP STYLE CONSTANTS
# ══════════════════════════════════════════════

XP = {
    'title_bg': '#C8102E',
    'title_fg': '#FFFFFF',
    'title_shadow': '#8B0000',
    'toolbar_bg': '#D4D0C8',
    'toolbar_border': '#808080',
    'bg': '#D4D0C8',
    'panel_bg': '#ECE9D8',
    'white': '#FFFFFF',
    'black': '#000000',
    'text': '#000000',
    'text2': '#444444',
    'text3': '#808080',
    'btn_face': '#ECE9D8',
    'btn_highlight': '#FFFFFF',
    'btn_shadow': '#808080',
    'btn_dark': '#404040',
    'btn_hover': '#F0EDE4',
    'border_light': '#FFFFFF',
    'border_dark': '#808080',
    'border_darker': '#404040',
    'field_bg': '#FFFFFF',
    'selection': '#C8102E',
    'selection_light': '#F4D4D9',
    'ok': '#008000',
    'err': '#CC0000',
    'warn': '#CC6600',
    'info': '#000080',
    'progress_bg': '#FFFFFF',
    'progress_fill': '#C8102E',
    'progress_chunk': '#E03040',
    'scrollbar': '#C0C0C0',
    'status_bg': '#ECE9D8',
    'status_border': '#808080',
    'terminal_bg': '#000000',
    'terminal_fg': '#C0C0C0',
}

# ══════════════════════════════════════════════
#  Helpers
# ══════════════════════════════════════════════

def parse_kibot_version(version_output):
    m = re.search(r'(\d+)\.(\d+)\.(\d+)', version_output or '')
    if not m:
        return None
    return tuple(int(x) for x in m.groups())

def is_kibot_version_compatible(version_output):
    parsed = parse_kibot_version(version_output)
    if not parsed:
        return False
    return parsed >= KIBOT_MIN_VERSION

def parse_kicad_major(version_str):
    m = re.search(r'(\d+)', version_str or '')
    return int(m.group(1)) if m else 0

def detect_kicad_version(filepath):
    try:
        with open(filepath, 'r', encoding='utf-8', errors='replace') as f: head = f.read(2000)
        m = re.search(r'\(version\s+(\d+)\)', head)
        if not m: return None, "unknown"
        ver = int(m.group(1))
        if filepath.endswith('.kicad_sch'):
            if ver > 20250114: return ver, "KiCad 10"
            if ver > 20231120: return ver, "KiCad 9"
            return ver, "KiCad 7/8"
        if filepath.endswith('.kicad_pcb'):
            if ver > 20241229: return ver, "KiCad 10"
            if ver > 20240108: return ver, "KiCad 9"
            return ver, "KiCad 7/8"
        return ver, "unknown"
    except: return None, "error"

def get_project_name(d):
    schs = glob.glob(os.path.join(d, '*.kicad_sch'))
    if not schs: return None
    for f in schs:
        nm = Path(f).stem
        if '_sheet' not in nm and '-sheet' not in nm and '_sub' not in nm: return nm
    return Path(schs[0]).stem

def wsl_path(win_path):
    if not IS_WINDOWS: return win_path
    p = os.path.abspath(win_path).replace('\\', '/')
    if len(p) >= 2 and p[1] == ':':
        p = '/mnt/' + p[0].lower() + p[2:]
    return p

def check_requirements():
    results = []
    if not IS_WINDOWS:
        results.append(("WSL", True, "Linux nativo"))
        try:
            v = subprocess.check_output(['kibot','--version'], text=True, stderr=subprocess.STDOUT).strip()
            ok = is_kibot_version_compatible(v)
            min_v = '.'.join(str(x) for x in KIBOT_MIN_VERSION)
            detail = v if ok else f"{v} (requiere >= {min_v})"
            results.append(("KiBot", ok, detail))
        except:
            results.append(("KiBot", False, "No encontrado"))
        try:
            kv = subprocess.check_output(
                ['python3','-c','import pcbnew; print(pcbnew.GetBuildVersion())'],
                text=True, stderr=subprocess.STDOUT).strip()
            ok = parse_kicad_major(kv) >= KICAD_MIN_MAJOR
            detail = kv if ok else f"{kv} (requiere KiCad {KICAD_MIN_MAJOR})"
            results.append(("KiCad", ok, detail))
        except:
            results.append(("KiCad", False, "No instalado"))
        return results
    try:
        subprocess.check_output(['wsl','echo','ok'], text=True, **POPEN_FLAGS)
        results.append(("WSL", True, "OK"))
    except:
        results.append(("WSL", False, "No instalado")); return results
    try:
        v = subprocess.check_output(
            ['wsl','bash','-lc','PATH="$HOME/.local/bin:$PATH" kibot --version'],
            text=True, encoding='utf-8', **POPEN_FLAGS).strip()
        ok = is_kibot_version_compatible(v)
        min_v = '.'.join(str(x) for x in KIBOT_MIN_VERSION)
        detail = v if ok else f"{v} (requiere >= {min_v})"
        results.append(("KiBot", ok, detail))
    except:
        results.append(("KiBot", False, "No instalado en WSL"))
    try:
        kv = subprocess.check_output(
            ['wsl','bash','-lc','python3 -c "import pcbnew; print(pcbnew.GetBuildVersion())"'],
            text=True, encoding='utf-8', **POPEN_FLAGS).strip()
        ok = parse_kicad_major(kv) >= KICAD_MIN_MAJOR
        detail = kv if ok else f"{kv} (requiere KiCad {KICAD_MIN_MAJOR})"
        results.append(("KiCad", ok, detail))
    except:
        results.append(("KiCad", False, "No instalado en WSL"))
    return results

def load_config():
    try:
        with open(CONFIG_FILE,'r') as f: return json.load(f)
    except: return {}

def save_config(c):
    try:
        with open(CONFIG_FILE,'w') as f: json.dump(c, f)
    except: pass

def validate_yaml(data):
    errs = []
    if not isinstance(data, dict): return False, ["YAML no es un diccionario"]
    if 'variants' not in data: return False, ["Falta clave 'variants'"]
    v = data['variants']
    if not isinstance(v, list): return False, ["'variants' debe ser una lista"]
    if len(v) == 0: return False, ["Lista 'variants' vacia"]
    for i, x in enumerate(v):
        if not isinstance(x, dict): errs.append(f"Variante {i+1}: no es diccionario")
        elif 'name' not in x: errs.append(f"Variante {i+1}: falta 'name'")
    return len(errs)==0, errs

def generate_pnp_chm551(pos_dir, bom_xlsx, log_cb):
    """Generate PnP CSVs for Charmhigh CHM-551 from position CSVs and BOM xlsx."""
    wb = load_workbook(bom_xlsx, data_only=True)
    ws = wb.active
    ref_to_torsa = {}
    for row in ws.iter_rows(min_row=2):
        references = row[1].value  # col B
        torsa = row[6].value       # col G
        if references and torsa:
            for ref in str(references).replace(",", " ").split():
                ref_to_torsa[ref.strip()] = str(torsa).strip()
    wb.close()
    log_cb(f"  BOM: {len(ref_to_torsa)} referencias cargadas")

    created = []
    for csv_in in sorted(glob.glob(os.path.join(pos_dir, '*_pos.csv'))):
        bn = os.path.basename(csv_in)
        if '_PnP_CHM551' in bn:
            continue
        is_bottom = '_bottom_pos.csv' in bn
        out_name = bn.replace('_pos.csv', '_pos_PnP_CHM551.csv')
        csv_out = os.path.join(pos_dir, out_name)

        rows = []
        fieldnames = None
        with open(csv_in, newline='', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            fieldnames = reader.fieldnames
            for r in reader:
                ref = r.get('Ref', '').strip()
                torsa = ref_to_torsa.get(ref)
                if torsa:
                    r['Val'] = torsa
                if is_bottom:
                    if 'PosX' in r:
                        try:
                            val = float(r['PosX'])
                            if val > 0:
                                r['PosX'] = f"{-val:.4f}"
                        except ValueError:
                            pass
                    if 'Rot' in r:
                        try:
                            a = (180.0 - float(r['Rot'])) % 360.0
                            if a > 180.0:
                                a -= 360.0
                            r['Rot'] = f"{a:.4f}"
                        except ValueError:
                            pass
                rows.append(r)

        with open(csv_out, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(rows)
        side = "BOTTOM (PosX negado)" if is_bottom else "TOP"
        log_cb(f"  {out_name}  [{side}]")
        created.append(csv_out)
    return created

def generate_odoo_bom_csv(bom_xlsx_path, code, log_cb):
    """Generate Odoo BOM import CSV from KiBot BOM xlsx."""
    wb = load_workbook(bom_xlsx_path, data_only=True)
    ws = wb.active
    product_tmpl_id = str(ws['D3'].value or '').strip()
    lines = []
    for row_num in range(9, ws.max_row + 1):
        product_id = ws.cell(row=row_num, column=7).value  # Column G (Torsa#)
        product_qty = ws.cell(row=row_num, column=3).value  # Column C (Quantity per PCB)
        if not product_id:
            break
        try:
            qty_f = float(product_qty)
            qty = int(qty_f) if qty_f == int(qty_f) else qty_f
        except (TypeError, ValueError):
            qty = product_qty
        lines.append((str(product_id).strip(), qty))
    wb.close()
    log_cb(f"  Producto: {product_tmpl_id}")
    log_cb(f"  Componentes: {len(lines)}")
    if not lines:
        log_cb("  Sin componentes en el BOM")
        return None
    bn = os.path.basename(bom_xlsx_path)
    csv_name = bn.replace('-bom.xlsx', '-odoo_bom.csv') if '-bom.xlsx' in bn else os.path.splitext(bn)[0] + '_odoo.csv'
    csv_path = os.path.join(os.path.dirname(bom_xlsx_path), csv_name)
    with open(csv_path, 'w', newline='', encoding='utf-8-sig') as f:
        w = csv.writer(f)
        w.writerow(['product_tmpl_id', 'code', 'type',
                    'bom_line_ids/product_id', 'bom_line_ids/product_qty', 'bom_line_ids/product_uom_id'])
        w.writerow([product_tmpl_id, code, 'Fabricar este producto',
                    lines[0][0], lines[0][1], 'Unidades'])
        for pid, qty in lines[1:]:
            w.writerow(['', '', '', pid, qty, 'Unidades'])
    log_cb(f"  Guardado: {csv_name}")
    return csv_path


# ══════════════════════════════════════════════
#  XP-STYLE WIDGETS
# ══════════════════════════════════════════════

class XPButton(tk.Button):
    """Styled button like Windows XP."""
    def __init__(self, parent, text="", command=None, width=120, height=28, **kw):
        self._w_px = width
        self._h_px = height
        super().__init__(parent, text=text, command=command,
                         font=("Tahoma", 9), relief='raised', bd=2,
                         bg=XP['btn_face'], fg=XP['text'],
                         activebackground='#D0CEC6', activeforeground=XP['text'],
                         cursor='hand2', padx=8, pady=4)
        self.bind('<Enter>', lambda e: self.config(bg=XP['btn_hover']) if self['state'] != 'disabled' else None)
        self.bind('<Leave>', lambda e: self.config(bg=XP['btn_face']) if self['state'] != 'disabled' else None)

    def set_state(self, enabled):
        self.config(state='normal' if enabled else 'disabled',
                    bg=XP['btn_face'] if enabled else '#C0C0C0',
                    cursor='hand2' if enabled else 'arrow')

    def set_text(self, t):
        self.config(text=t)


class XPProgress(tk.Frame):
    """Chunked progress bar like Windows XP."""
    def __init__(self, parent, width=300, height=22, **kw):
        super().__init__(parent, height=height, bg=XP['bg'])
        self._canvas = tk.Canvas(self, width=width, height=height,
                                  highlightthickness=0, bg=XP['bg'])
        self._canvas.pack(fill='x', expand=True)
        self._h = height
        self._value = 0
        self.bind('<Configure>', lambda e: self._draw())

    def _draw(self):
        c = self._canvas
        c.delete('all')
        w = c.winfo_width() or 300
        h = self._h
        c.create_rectangle(0, 0, w-1, h-1, outline=XP['btn_shadow'])
        c.create_rectangle(1, 1, w-2, h-2, outline=XP['btn_dark'])
        c.create_rectangle(2, 2, w-3, h-3, fill=XP['progress_bg'], outline='')
        if self._value > 0:
            fill_w = int((w - 6) * min(self._value, 1.0))
            x = 3
            while x < 3 + fill_w:
                cw = min(8, 3 + fill_w - x)
                c.create_rectangle(x, 3, x+cw, h-4, fill=XP['progress_fill'], outline='')
                c.create_rectangle(x, 3, x+cw, 6, fill=XP['progress_chunk'], outline='')
                x += 10

    def set_value(self, v):
        self._value = v; self._draw()


# ══════════════════════════════════════════════
#  GUI
# ══════════════════════════════════════════════

class KiBotGUI(TkinterDnD.Tk if HAS_DND else tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("KiBot Variant Manager")
        self.geometry("1050x720")
        self.minsize(860, 560)
        self.configure(bg=XP['bg'])

        # ── load images ──
        self._images = {}
        try:
            logo = Image.open(os.path.join(ASSETS_DIR, 'Logo_Torsa.png'))
            logo = logo.resize((120, int(120 * logo.height / logo.width)), Image.LANCZOS)
            self._images['logo'] = ImageTk.PhotoImage(logo)
        except Exception:
            self._images['logo'] = None
        try:
            icn = Image.open(os.path.join(ASSETS_DIR, 'icono.png'))
            self._images['icon_drop'] = ImageTk.PhotoImage(icn.resize((80, 80), Image.LANCZOS))
            ico16 = ImageTk.PhotoImage(icn.resize((16, 16), Image.LANCZOS))
            ico32 = ImageTk.PhotoImage(icn.resize((32, 32), Image.LANCZOS))
            ico48 = ImageTk.PhotoImage(icn.resize((48, 48), Image.LANCZOS))
            self.iconphoto(True, ico48, ico32, ico16)
        except Exception:
            self._images['icon_drop'] = None

        self.variants = []
        self.yaml_path = None
        self.project_dir = None
        self.current_proc = None
        self.requirements_ok = False
        self.execution_history = []
        self.variant_buttons = {}
        self.detected_version = ""
        self._yaml_data = {}

        self._build_ui()
        self._check_startup()

    def _build_ui(self):
        # ═══ TITLE BAR ═══
        title = tk.Frame(self, bg=XP['title_bg'], height=36)
        title.pack(fill='x')
        title.pack_propagate(False)
        # shadow text
        tk.Label(title, text=" KiBot Variant Manager", fg=XP['title_shadow'],
                 bg=XP['title_bg'], font=("Tahoma", 12, "bold")).place(x=7, y=7)
        tk.Label(title, text=" KiBot Variant Manager", fg=XP['title_fg'],
                 bg=XP['title_bg'], font=("Tahoma", 12, "bold")).place(x=6, y=6)
        if self._images.get('logo'):
            tk.Label(title, image=self._images['logo'], bg=XP['title_bg']).pack(side='right', padx=8)

        # ═══ TOOLBAR ═══
        toolbar = tk.Frame(self, bg=XP['toolbar_bg'], height=38)
        toolbar.pack(fill='x')
        toolbar.pack_propagate(False)
        # separator
        tk.Frame(toolbar, bg=XP['btn_shadow'], height=1).pack(fill='x', side='top')

        tbtn_frame = tk.Frame(toolbar, bg=XP['toolbar_bg'])
        tbtn_frame.pack(fill='x', padx=4, pady=3)

        self.btn_load = XPButton(tbtn_frame, text="Cargar YAML", command=self._load_yaml, width=110, height=26)
        self.btn_load.pack(side='left', padx=2)
        self.btn_clear = XPButton(tbtn_frame, text="Limpiar", command=self._clear_all, width=80, height=26)
        self.btn_clear.pack(side='left', padx=2)
        self.btn_cancel = XPButton(tbtn_frame, text="Cancelar", command=self._cancel_process, width=80, height=26)
        self.btn_cancel.pack(side='left', padx=2)
        self.btn_cancel.set_state(False)

        tk.Frame(tbtn_frame, bg=XP['btn_shadow'], width=1).pack(side='left', fill='y', padx=4)
        self.btn_open_output = XPButton(tbtn_frame, text="Abrir salida", command=self._open_output_folder, width=95, height=26)
        self.btn_open_output.pack(side='left', padx=2)
        self.btn_open_output.set_state(False)
        self.btn_export_log = XPButton(tbtn_frame, text="Exportar log", command=self._export_log, width=90, height=26)
        self.btn_export_log.pack(side='left', padx=2)
        self.btn_export_log.set_state(False)

        tk.Frame(self, bg=XP['btn_shadow'], height=1).pack(fill='x')

        # ═══ DROP ZONE ═══
        self.drop_frame = tk.Frame(self, bg=XP['bg'])
        self.drop_frame.pack(fill='both', expand=True)
        self._build_drop_zone()

        # ═══ MAIN FRAME ═══
        self.main_frame = tk.Frame(self, bg=XP['bg'])

        # ── LEFT SIDEBAR ──
        sidebar = tk.Frame(self.main_frame, bg=XP['panel_bg'], width=220)
        sidebar.pack(side='left', fill='y', padx=(4,0), pady=4)
        sidebar.pack_propagate(False)

        # sidebar border
        tk.Frame(sidebar, bg=XP['btn_shadow'], width=1).pack(side='right', fill='y')

        # sidebar header
        sh = tk.Frame(sidebar, bg=XP['title_bg'], height=24)
        sh.pack(fill='x')
        sh.pack_propagate(False)
        tk.Label(sh, text="  Variantes", fg='white', bg=XP['title_bg'],
                 font=("Tahoma", 9, "bold")).pack(side='left')

        self.btn_frame = tk.Frame(sidebar, bg=XP['panel_bg'])
        self.btn_frame.pack(fill='both', expand=True, padx=4, pady=4)

        # bottom sidebar buttons
        sidebar_bottom = tk.Frame(sidebar, bg=XP['panel_bg'])
        sidebar_bottom.pack(side='bottom', fill='x', padx=4, pady=4)
        self.btn_run_all = XPButton(sidebar_bottom, text="Ejecutar todas", command=self._run_all_variants, width=200, height=28)
        self.btn_run_all.pack(fill='x', pady=(0,2))
        self.btn_run_all.set_state(False)

        # ── RIGHT AREA ──
        right = tk.Frame(self.main_frame, bg=XP['bg'])
        right.pack(side='left', fill='both', expand=True, padx=4, pady=4)

        # ── PROJECT CARD ──
        self.project_card = tk.LabelFrame(right, text="  Proyecto  ", bg=XP['panel_bg'],
                                           fg=XP['info'], font=("Tahoma", 9, "bold"),
                                           relief='groove', bd=2, padx=8, pady=4)
        self.card_vars = {}
        for _key, _lbl in [('nombre', 'PCB'), ('variantes', 'Variantes'), ('version', 'Version KiCad'),
                             ('fecha', 'Fecha YAML'), ('outputs', 'Outputs'), ('ruta', 'Ruta')]:
            _row = tk.Frame(self.project_card, bg=XP['panel_bg'])
            _row.pack(fill='x', pady=0)
            tk.Label(_row, text=f"{_lbl}:", bg=XP['panel_bg'], fg=XP['text2'],
                     font=("Tahoma", 8, "bold"), width=12, anchor='w').pack(side='left')
            _v = tk.StringVar()
            self.card_vars[_key] = _v
            tk.Label(_row, textvariable=_v, bg=XP['panel_bg'], fg=XP['text'],
                     font=("Tahoma", 8), anchor='w').pack(side='left', fill='x', expand=True)

        # ── STATUS BOX (groupbox style) ──
        self.status_outer = tk.Frame(right, bg=XP['bg'])
        self.status_outer.pack(fill='x', pady=(0,4))
        status_outer = self.status_outer

        # groupbox border
        gb = tk.LabelFrame(status_outer, text="  Estado  ", bg=XP['panel_bg'],
                           fg=XP['info'], font=("Tahoma", 9, "bold"),
                           relief='groove', bd=2, padx=8, pady=4)
        gb.pack(fill='x')

        self.steps_frame = tk.Frame(gb, bg=XP['panel_bg'])
        self.steps_frame.pack(fill='x')

        self.step_labels = {}
        for sid, stxt in [('yaml','YAML cargado'), ('version','Version compatible'), ('ready','Listo para generar')]:
            row = tk.Frame(self.steps_frame, bg=XP['panel_bg'])
            row.pack(fill='x', pady=1)
            ind = tk.Label(row, text="[ -- ]", bg=XP['panel_bg'], fg=XP['text3'],
                           font=("Tahoma", 8, "bold"), width=8, anchor='w')
            ind.pack(side='left')
            lbl = tk.Label(row, text=stxt, bg=XP['panel_bg'], fg=XP['text3'],
                           font=("Tahoma", 9), anchor='w')
            lbl.pack(side='left')
            self.step_labels[sid] = (ind, lbl)

        # ── PROGRESS BAR ──
        prog_frame = tk.LabelFrame(right, text="  Progreso  ", bg=XP['panel_bg'],
                                    fg=XP['info'], font=("Tahoma", 9, "bold"),
                                    relief='groove', bd=2, padx=8, pady=6)
        prog_frame.pack(fill='x', pady=(0,4))

        self.progress = XPProgress(prog_frame, width=600, height=22)
        self.progress.pack(fill='x')

        # ── HISTORIAL ──
        hist_gb = tk.LabelFrame(right, text="  Historial de ejecuciones  ", bg=XP['panel_bg'],
                                 fg=XP['info'], font=("Tahoma", 9, "bold"),
                                 relief='groove', bd=2, padx=4, pady=4)
        hist_gb.pack(fill='x', pady=(0,4))
        self.hist_list = tk.Listbox(hist_gb, bg=XP['white'], fg=XP['text'],
                                     font=("Lucida Console", 8) if IS_WINDOWS else ("Consolas", 8),
                                     height=4, relief='sunken', bd=2,
                                     selectbackground=XP['selection'], selectforeground='white')
        self.hist_list.pack(fill='x')
        self.hist_list.bind('<Double-Button-1>', self._on_hist_dblclick)
        self._hist_variants = []

        # ── TERMINAL ──
        term_gb = tk.LabelFrame(right, text="  Terminal  ", bg=XP['bg'],
                                 fg=XP['info'], font=("Tahoma", 9, "bold"),
                                 relief='groove', bd=2, padx=4, pady=4)
        term_gb.pack(fill='both', expand=True)

        # sunken frame for terminal
        term_border = tk.Frame(term_gb, bg=XP['btn_shadow'], bd=0)
        term_border.pack(fill='both', expand=True)

        self.log = scrolledtext.ScrolledText(term_border, bg=XP['terminal_bg'], fg=XP['terminal_fg'],
                                              font=("Lucida Console", 9) if IS_WINDOWS else ("Consolas", 9),
                                              wrap='word', insertbackground='white',
                                              relief='sunken', bd=2)
        self.log.pack(fill='both', expand=True, padx=1, pady=1)
        self.log.tag_config('err', foreground='#FF6B6B')
        self.log.tag_config('warn', foreground='#FFB347')
        self.log.tag_config('ok', foreground='#90EE90')
        self.log.tag_config('info', foreground='#87CEEB')
        self.log.tag_config('dim', foreground='#666666')

        # ═══ STATUS BAR ═══
        sb_frame = tk.Frame(self, bg=XP['status_bg'], height=24)
        sb_frame.pack(side='bottom', fill='x')
        sb_frame.pack_propagate(False)
        # top border
        tk.Frame(sb_frame, bg=XP['btn_shadow'], height=1).pack(fill='x', side='top')
        # grip
        tk.Frame(sb_frame, bg=XP['btn_shadow'], width=2).pack(side='left', fill='y', padx=(2,4))

        self.status_var = tk.StringVar(value="Iniciando...")
        tk.Label(sb_frame, textvariable=self.status_var, bg=XP['status_bg'], fg=XP['text2'],
                 font=("Tahoma", 8), anchor='w').pack(side='left', fill='x', expand=True, padx=4)

    def _build_drop_zone(self):
        # XP explorer-like window area
        outer = tk.Frame(self.drop_frame, bg=XP['panel_bg'], relief='sunken', bd=2)
        outer.pack(expand=True, fill='both', padx=20, pady=20)

        inner = tk.Frame(outer, bg=XP['white'])
        inner.pack(expand=True, fill='both', padx=2, pady=2)

        center = tk.Frame(inner, bg=XP['white'])
        center.place(relx=0.5, rely=0.4, anchor='center')

        # app icon
        if self._images.get('icon_drop'):
            tk.Label(center, image=self._images['icon_drop'], bg=XP['white']).pack()
        else:
            tk.Label(center, text="[  ]", font=("Tahoma", 48, "bold"),
                     bg=XP['white'], fg='#FFD700').pack()
        tk.Label(center, text="Arrastre su archivo YAML aqui",
                 font=("Tahoma", 13, "bold"), bg=XP['white'], fg=XP['text']).pack(pady=(8,2))
        tk.Label(center, text="o use el boton 'Cargar YAML' de la barra de herramientas",
                 font=("Tahoma", 9), bg=XP['white'], fg=XP['text3']).pack()

        if HAS_DND:
            inner.drop_target_register(DND_FILES)
            inner.dnd_bind('<<DropEnter>>', lambda e: inner.configure(bg=XP['white']))
            inner.dnd_bind('<<DropLeave>>', lambda e: inner.configure(bg=XP['white']))
            inner.dnd_bind('<<Drop>>', self._on_drop)
            tk.Label(center, text="[Drag & Drop activo]", font=("Tahoma", 8, "bold"),
                     bg=XP['white'], fg=XP['ok']).pack(pady=(10,0))
        else:
            tk.Label(center, text="Instale tkinterdnd2 para arrastrar y soltar",
                     font=("Tahoma", 8), bg=XP['white'], fg=XP['warn']).pack(pady=(10,0))

        # Requirements at bottom
        self.req_frame = tk.LabelFrame(inner, text="  Requisitos del sistema  ",
                                        bg=XP['white'], fg=XP['info'],
                                        font=("Tahoma", 8, "bold"), relief='groove', bd=2)
        self.req_frame.pack(side='bottom', fill='x', padx=16, pady=12)

    # ── STARTUP ──

    def _check_startup(self):
        self.status_var.set("Comprobando requisitos...")
        self.update_idletasks()
        reqs = check_requirements()
        self.requirements_ok = all(ok for _, ok, _ in reqs)
        for w in self.req_frame.winfo_children(): w.destroy()
        for name, ok, detail in reqs:
            row = tk.Frame(self.req_frame, bg=XP['white'])
            row.pack(fill='x', pady=1, padx=4)
            icon = "[OK]" if ok else "[!!]"
            color = XP['ok'] if ok else XP['err']
            tk.Label(row, text=icon, bg=XP['white'], fg=color, font=("Tahoma", 8, "bold"),
                     width=5, anchor='w').pack(side='left')
            tk.Label(row, text=f"{name}: {detail}", bg=XP['white'], fg=XP['text2'],
                     font=("Tahoma", 8), anchor='w').pack(side='left')
        self.status_var.set("Listo" if self.requirements_ok else "Faltan requisitos del sistema")

    # ── STEP INDICATORS ──

    def _set_step(self, sid, state):
        ind, lbl = self.step_labels[sid]
        m = {'ok':('[OK]', XP['ok']), 'error':('[!!]', XP['err']),
             'running':('[>>]', XP['info']), 'pending':('[ -- ]', XP['text3'])}
        txt, col = m.get(state, ('[ -- ]', XP['text3']))
        ind.config(text=txt, fg=col); lbl.config(fg=col)

    # ── EVENTS ──

    def _on_drop(self, event):
        raw = event.data
        paths = re.findall(r'\{([^}]+)\}', raw) if '{' in raw else raw.split()
        for p in paths:
            p = p.strip()
            if p.lower().endswith(('.yaml','.yml')) and os.path.isfile(p):
                self._process_yaml(p); return
        messagebox.showwarning("Archivo", "Suelte un archivo .yaml o .yml")

    def _log(self, msg, tag=None):
        self.log.insert('end', msg + '\n', tag or ''); self.log.see('end')

    def _clear_all(self):
        self.variants = []; self.yaml_path = None; self.project_dir = None
        self.variant_buttons = {}; self._yaml_data = {}
        for w in self.btn_frame.winfo_children(): w.destroy()
        self.log.delete('1.0', 'end')
        self.hist_list.delete(0, 'end')
        self._hist_variants = []
        self.project_card.pack_forget()
        for s in self.step_labels: self._set_step(s, 'pending')
        self.progress.set_value(0)
        self.btn_open_output.set_state(False)
        self.btn_export_log.set_state(False)
        self.btn_run_all.set_state(False)
        self.main_frame.pack_forget()
        self.drop_frame.pack(fill='both', expand=True)
        self.status_var.set("Listo")

    def _load_yaml(self):
        cfg = load_config()
        ini = os.path.dirname(cfg.get('last_yaml','')) or None
        path = filedialog.askopenfilename(title="Seleccionar YAML KiBot",
                                           filetypes=[("YAML","*.yaml *.yml"),("Todos","*.*")],
                                           initialdir=ini)
        if path: self._process_yaml(path)

    def _process_yaml(self, path):
        if not self.requirements_ok:
            messagebox.showerror("Requisitos", "Faltan requisitos.\nRevise WSL, KiCad y KiBot >= 1.9.0."); return
        self.yaml_path = path; self.project_dir = os.path.dirname(path)
        save_config({'last_yaml': path})
        try:
            with open(path, 'r', encoding='utf-8') as f: data = yaml.safe_load(f)
        except Exception as e:
            messagebox.showerror("Error YAML", str(e)); return
        ok, errs = validate_yaml(data)
        if not ok:
            self._set_step('yaml','error')
            messagebox.showerror("YAML invalido", "\n".join(errs)); return
        self.variants = data['variants']
        self._yaml_data = data
        self.drop_frame.pack_forget(); self.main_frame.pack(fill='both', expand=True)
        self.log.delete('1.0','end')
        self._log(f"YAML: {os.path.basename(path)}", 'info')
        self._log(f"Ruta: {self.project_dir}", 'dim')
        self._log(f"Variantes: {len(self.variants)}", 'info')
        self._log("", None)
        self._set_step('yaml','ok')
        self.status_var.set(f"{os.path.basename(self.project_dir)} | {len(self.variants)} variantes")
        self.btn_open_output.set_state(True)
        self.btn_export_log.set_state(True)
        self.variant_buttons = {}
        for w in self.btn_frame.winfo_children(): w.destroy()
        for var in self.variants:
            nm = var.get('name','???'); cm = var.get('comment','')
            btn = XPButton(self.btn_frame, text=nm, command=lambda v=var: self._run_variant(v),
                           width=200, height=26)
            btn.pack(fill='x', pady=2)
            self.variant_buttons[nm] = btn
            if cm:
                tk.Label(self.btn_frame, text=f"  {cm}", bg=XP['panel_bg'], fg=XP['text3'],
                         font=("Tahoma", 7), anchor='w').pack(fill='x')
        self._check_versions()
        self._validate_yaml_refs()
        self._update_project_card()

    def _check_versions(self):
        self.detected_version = ""
        if not self.project_dir: return
        files = glob.glob(os.path.join(self.project_dir,'*.kicad_sch')) + \
                glob.glob(os.path.join(self.project_dir,'*.kicad_pcb'))
        if not files:
            self._set_step('version','error'); self._set_step('ready','pending')
            self._log("No se encontraron archivos KiCad (.kicad_sch/.kicad_pcb).\n", 'err')
            return

        self._log("Comprobando versiones...", 'dim')
        labels = []
        for fp in files:
            _, label = detect_kicad_version(fp)
            if label not in labels:
                labels.append(label)
            tag = 'ok' if label in ("KiCad 7/8", "KiCad 9", "KiCad 10") else 'warn'
            self._log(f"  {os.path.basename(fp)}: {label}", tag)
        self.detected_version = ", ".join(labels)
        self._set_step('version','ok'); self._set_step('ready','ok')
        self._log("Versiones compatibles con KiBot 1.9.0.\n", 'ok')
        for w in self.btn_frame.winfo_children():
            if isinstance(w, XPButton): w.set_state(True)
        self.btn_run_all.set_state(True)

    def _validate_yaml_refs(self):
        """Check that files referenced in the YAML (logo, etc.) exist in the project dir."""
        data = self._yaml_data
        refs_found = []

        def _scan(obj):
            if isinstance(obj, str):
                # Detect file references: strings ending with image/file extensions
                if re.match(r'^[\w\-. ]+\.(png|jpg|svg|step|stp|wrl|kicad_dbl)$', obj, re.IGNORECASE):
                    refs_found.append(obj)
            elif isinstance(obj, dict):
                for v in obj.values(): _scan(v)
            elif isinstance(obj, list):
                for v in obj: _scan(v)

        _scan(data)
        if not refs_found:
            return
        self._log("Comprobando archivos referenciados...", 'dim')
        all_ok = True
        for ref in sorted(set(refs_found)):
            fp = os.path.join(self.project_dir, ref)
            if os.path.isfile(fp):
                self._log(f"  {ref}: encontrado", 'ok')
            else:
                self._log(f"  {ref}: NO encontrado", 'warn')
                all_ok = False
        if not all_ok:
            self._log("  Algunos archivos referenciados no estan en la carpeta del proyecto.", 'warn')
            self._log("  KiBot podria generar warnings o errores.\n", 'warn')

    # ── RUN VARIANT ──

    def _run_variant(self, var, on_done=None):
        # Validate KiCad files exist
        schs = glob.glob(os.path.join(self.project_dir, '*.kicad_sch'))
        pcbs = glob.glob(os.path.join(self.project_dir, '*.kicad_pcb'))
        if not schs or not pcbs:
            missing = []
            if not schs: missing.append('.kicad_sch')
            if not pcbs: missing.append('.kicad_pcb')
            messagebox.showerror("Archivos faltantes",
                f"No se encontraron archivos {' ni '.join(missing)}\n"
                f"en {self.project_dir}")
            if on_done: self.after(0, on_done, var.get('name',''), False)
            return
        name = var.get('name','')
        self._log(f"\n{'='*50}", 'dim')
        self._log(f"Ejecutando: {name}", 'info')
        self._set_step('ready','running'); self.progress.set_value(0.1)
        self.btn_cancel.set_state(True)
        self.status_var.set(f"Ejecutando: {name}...")
        # Count expected outputs for real progress
        self._output_total = len(self._yaml_data.get('outputs', []))
        self._output_count = 0

        if IS_WINDOWS:
            wy = wsl_path(self.yaml_path); wd = wsl_path(self.project_dir)
            full = f'cd "{wd}" && PATH="$HOME/.local/bin:$PATH" kibot -c "{wy}" -g variant={name}'
            cmd = ['wsl', 'bash', '-lc', full]
        else:
            cmd = ['kibot', '-c', self.yaml_path, '-g', f'variant={name}']
        self._log(f"$ {' '.join(cmd)}", 'dim')

        def run():
            success = False
            try:
                self.current_proc = subprocess.Popen(cmd, cwd=None if IS_WINDOWS else self.project_dir,
                    stdout=subprocess.PIPE, stderr=subprocess.STDOUT, text=True, encoding='utf-8', **POPEN_FLAGS)
                for line in self.current_proc.stdout:
                    self.after(0, self._log_and_track_progress, line.rstrip())
                self.current_proc.wait(); rc = self.current_proc.returncode; self.current_proc = None
                if rc == 0:
                    success = True
                    self.after(0, self._set_step, 'ready', 'ok')
                    self.after(0, self.progress.set_value, 1.0)
                    self.after(0, self._log, "Finalizado OK", 'ok')
                    self.after(0, self.status_var.set, f"{name} completado")
                    self.after(0, self._mark_variant, name, 'ok')
                    if not on_done:
                        self.after(100, self._ask_pnp, name)
                elif rc in (-9, 137):
                    self.after(0, self._log, "Cancelado.", 'warn')
                    self.after(0, self._set_step, 'ready', 'ok')
                    self.after(0, self.status_var.set, "Cancelado")
                else:
                    self.after(0, self._set_step, 'ready', 'error')
                    self.after(0, self._log, f"Error (exit {rc})", 'err')
                    self.after(0, self.status_var.set, f"{name} fallo")
                    self.after(0, self._mark_variant, name, 'err')
            except FileNotFoundError:
                self.after(0, self._log, "ERROR: comando no encontrado", 'err')
            except Exception as e:
                self.after(0, self._log, f"ERROR: {e}", 'err')
            finally:
                self.after(0, self.btn_cancel.set_state, False)
                self.after(0, self.progress.set_value, 0)
                if on_done:
                    self.after(0, on_done, name, success)
        threading.Thread(target=run, daemon=True).start()

    def _log_and_track_progress(self, line):
        self._log(line)
        # KiBot outputs "- 'Name' (type) [dir]" for each output it processes
        if line.startswith("- '") and self._output_total > 0:
            self._output_count += 1
            pct = self._output_count / self._output_total
            self.progress.set_value(min(pct, 0.95))

    def _cancel_process(self):
        if self.current_proc:
            try:
                if IS_WINDOWS:
                    # taskkill /T kills the entire process tree (wsl.exe + child kibot)
                    subprocess.run(['taskkill', '/F', '/T', '/PID', str(self.current_proc.pid)],
                                   timeout=5, **POPEN_FLAGS)
                else:
                    self.current_proc.kill()
                self._log("Cancelando...", 'warn')
            except Exception:
                try: self.current_proc.kill()
                except Exception: pass

    def _run_all_variants(self):
        self._run_queue = list(self.variants)
        self._run_all_total = len(self._run_queue)
        self._run_all_ok = 0
        self._log(f"\n{'='*50}", 'dim')
        self._log(f"Ejecutando todas las variantes ({self._run_all_total})...", 'info')
        self.btn_run_all.set_state(False)
        self._run_next_in_queue()

    def _run_next_in_queue(self):
        if not self._run_queue:
            self._log(f"\nTodas completadas: {self._run_all_ok}/{self._run_all_total}", 'ok')
            self.status_var.set(f"Completadas: {self._run_all_ok}/{self._run_all_total}")
            self.btn_run_all.set_state(True)
            return
        var = self._run_queue.pop(0)
        self._run_variant(var, on_done=self._on_queue_variant_done)

    def _on_queue_variant_done(self, name, success):
        if success:
            self._run_all_ok += 1
        remaining = len(self._run_queue)
        self.status_var.set(f"Cola: {self._run_all_total - remaining}/{self._run_all_total}")
        if remaining > 0:
            self.after(200, self._run_next_in_queue)
        else:
            self._log(f"\nTodas completadas: {self._run_all_ok}/{self._run_all_total}", 'ok')
            self.status_var.set(f"Completadas: {self._run_all_ok}/{self._run_all_total}")
            self.btn_run_all.set_state(True)

    # ── NEW FEATURES ──

    def _update_project_card(self):
        path = self.yaml_path
        data = self._yaml_data
        nombre = get_project_name(self.project_dir) or Path(path).stem
        fecha = datetime.datetime.fromtimestamp(os.path.getmtime(path)).strftime("%d/%m/%Y %H:%M")
        outputs = data.get('outputs', [])
        out_types = []
        for o in outputs:
            if isinstance(o, dict):
                ot = o.get('type', o.get('name', '?'))
                if ot and ot not in out_types:
                    out_types.append(ot)
        out_str = ", ".join(out_types) if out_types else "No especificados"
        ruta = self.project_dir
        if len(ruta) > 55:
            ruta = "..." + ruta[-52:]
        self.card_vars['nombre'].set(nombre)
        self.card_vars['variantes'].set(str(len(self.variants)))
        self.card_vars['version'].set(self.detected_version or "Desconocida")
        self.card_vars['fecha'].set(fecha)
        self.card_vars['outputs'].set(out_str)
        self.card_vars['ruta'].set(ruta)
        self.project_card.pack(fill='x', pady=(0,4), before=self.status_outer)

    def _mark_variant(self, name, state):
        btn = self.variant_buttons.get(name)
        if btn:
            color = '#90EE90' if state == 'ok' else '#FF9999'
            btn.config(bg=color, activebackground=color)
            btn.bind('<Leave>', lambda _e, c=color: btn.config(bg=c))
        ts = datetime.datetime.now().strftime("%H:%M:%S")
        result = "  OK  " if state == 'ok' else "  FAIL"
        entry = f"[{ts}]  {name:<20}  {result}"
        self.hist_list.insert('end', entry)
        idx = self.hist_list.size() - 1
        self.hist_list.itemconfig(idx, fg=XP['ok'] if state == 'ok' else XP['err'])
        self.hist_list.see(idx)
        self._hist_variants.append(name)

    def _on_hist_dblclick(self, event):
        sel = self.hist_list.curselection()
        if not sel:
            return
        idx = sel[0]
        if idx >= len(self._hist_variants):
            return
        vname = self._hist_variants[idx]
        # Search for output folder containing this variant name
        for entry in os.scandir(self.project_dir):
            if entry.is_dir() and vname in entry.name:
                folder = entry.path
                if IS_WINDOWS:
                    os.startfile(folder)
                else:
                    subprocess.Popen(['xdg-open', folder])
                return
        # Fallback: open project dir
        if IS_WINDOWS:
            os.startfile(self.project_dir)
        else:
            subprocess.Popen(['xdg-open', self.project_dir])

    def _open_output_folder(self):
        folder = self.project_dir
        if folder and os.path.isdir(folder):
            if IS_WINDOWS:
                os.startfile(folder)
            else:
                subprocess.Popen(['xdg-open', folder])
        else:
            messagebox.showinfo("Carpeta", "Ruta de salida no disponible")

    def _export_log(self):
        path = filedialog.asksaveasfilename(
            title="Exportar log",
            defaultextension=".txt",
            filetypes=[("Texto", "*.txt"), ("Todos", "*.*")]
        )
        if path:
            content = self.log.get('1.0', 'end')
            with open(path, 'w', encoding='utf-8') as f:
                f.write(content)
            self.status_var.set(f"Log exportado: {os.path.basename(path)}")

    # ── PnP CHM-551 ──

    def _ask_pnp(self, variant_name):
        if not HAS_OPENPYXL:
            return
        win = tk.Toplevel(self)
        win.title("Pick & Place")
        win.geometry("420x200")
        win.configure(bg=XP['bg'])
        win.resizable(False, False)
        win.transient(self)
        win.grab_set()

        # XP title bar
        t = tk.Frame(win, bg=XP['title_bg'], height=28)
        t.pack(fill='x'); t.pack_propagate(False)
        tk.Label(t, text="  Pick & Place - CHM-551", fg='white', bg=XP['title_bg'],
                 font=("Tahoma", 10, "bold")).pack(side='left')

        body = tk.Frame(win, bg=XP['panel_bg'])
        body.pack(fill='both', expand=True, padx=8, pady=8)

        # Icon + message
        msg_frame = tk.Frame(body, bg=XP['panel_bg'])
        msg_frame.pack(fill='x', pady=(8, 12))
        tk.Label(msg_frame, text="[?]", bg=XP['panel_bg'], fg=XP['info'],
                 font=("Tahoma", 20, "bold")).pack(side='left', padx=(12, 8))
        msg_text = tk.Frame(msg_frame, bg=XP['panel_bg'])
        msg_text.pack(side='left', fill='x', expand=True)
        tk.Label(msg_text, text=f"Variante '{variant_name}' generada correctamente.",
                 bg=XP['panel_bg'], fg=XP['text'], font=("Tahoma", 9, "bold"),
                 anchor='w').pack(fill='x')
        tk.Label(msg_text, text="Desea crear los archivos PnP para\nCharmhigh CHM-551?",
                 bg=XP['panel_bg'], fg=XP['text'], font=("Tahoma", 9),
                 anchor='w', justify='left').pack(fill='x', pady=(4,0))

        # Buttons
        btn_frame = tk.Frame(body, bg=XP['panel_bg'])
        btn_frame.pack(pady=(0, 4))

        def on_yes():
            win.destroy()
            self._run_pnp(variant_name)
            self.after(150, self._ask_odoo_bom, variant_name)
        def on_no():
            win.destroy()
            self.after(150, self._ask_odoo_bom, variant_name)

        XPButton(btn_frame, text="Si, generar", command=on_yes, width=120, height=28).pack(side='left', padx=6)
        XPButton(btn_frame, text="No, finalizar", command=on_no, width=120, height=28).pack(side='left', padx=6)

        # Center on parent
        win.update_idletasks()
        x = self.winfo_x() + (self.winfo_width() - win.winfo_width()) // 2
        y = self.winfo_y() + (self.winfo_height() - win.winfo_height()) // 2
        win.geometry(f"+{x}+{y}")

    def _run_pnp(self, variant_name):
        self._log(f"\n{'='*50}", 'dim')
        self._log("Generando PnP para CHM-551...", 'info')
        self.status_var.set("Generando PnP CHM-551...")

        # Find Position folder: search for recently modified dirs matching this variant
        pos_dir = None
        bom_xlsx = None
        for root, dirs, files in os.walk(self.project_dir):
            bn = os.path.basename(root)
            if bn == 'Position':
                csvs = [f for f in files if f.endswith('_pos.csv') and '_PnP_CHM551' not in f]
                if csvs:
                    # Check sibling BoM folder for xlsx
                    parent = os.path.dirname(root)
                    bom_dir = os.path.join(parent, 'BoM')
                    xlsxs = glob.glob(os.path.join(bom_dir, '*.xlsx')) if os.path.isdir(bom_dir) else []
                    if xlsxs:
                        pos_dir = root
                        bom_xlsx = xlsxs[0]
                        break

        if not pos_dir or not bom_xlsx:
            self._log("No se encontraron carpetas Position/BoM con datos", 'err')
            self.status_var.set("PnP: carpetas no encontradas")
            return

        self._log(f"  Position: {pos_dir}", 'dim')
        self._log(f"  BOM: {os.path.basename(bom_xlsx)}", 'dim')

        try:
            created = generate_pnp_chm551(pos_dir, bom_xlsx, lambda m: self._log(m, 'ok'))
            self._log(f"PnP CHM-551: {len(created)} archivo(s) generados", 'ok')
            self.status_var.set(f"PnP CHM-551: {len(created)} archivo(s)")
        except Exception as e:
            self._log(f"Error generando PnP: {e}", 'err')
            self.status_var.set("PnP: error")

    # ── ODOO BOM CSV ──

    def _ask_odoo_bom(self, variant_name):
        if not HAS_OPENPYXL:
            return
        bom_xlsx = None
        for f in glob.glob(os.path.join(self.project_dir, '**', '*-bom.xlsx'), recursive=True):
            if variant_name in f:
                bom_xlsx = f
                break
        if not bom_xlsx:
            return

        win = tk.Toplevel(self)
        win.title("Exportar BOM para Odoo")
        win.geometry("460x210")
        win.configure(bg=XP['bg'])
        win.resizable(False, False)
        win.transient(self)
        win.grab_set()

        t = tk.Frame(win, bg=XP['title_bg'], height=28)
        t.pack(fill='x'); t.pack_propagate(False)
        tk.Label(t, text="  Exportar BOM para Odoo", fg='white', bg=XP['title_bg'],
                 font=("Tahoma", 10, "bold")).pack(side='left')

        body = tk.Frame(win, bg=XP['panel_bg'])
        body.pack(fill='both', expand=True, padx=8, pady=8)

        info = tk.Frame(body, bg=XP['panel_bg'])
        info.pack(fill='x', pady=(4, 8))
        tk.Label(info, text="[?]", bg=XP['panel_bg'], fg=XP['info'],
                 font=("Tahoma", 18, "bold")).pack(side='left', padx=(8, 10))
        msg = tk.Frame(info, bg=XP['panel_bg'])
        msg.pack(side='left', fill='x', expand=True)
        tk.Label(msg, text=f"BOM generado para variante '{variant_name}'.",
                 bg=XP['panel_bg'], fg=XP['text'], font=("Tahoma", 9, "bold"), anchor='w').pack(fill='x')
        tk.Label(msg, text=os.path.basename(bom_xlsx),
                 bg=XP['panel_bg'], fg=XP['text3'], font=("Tahoma", 8), anchor='w').pack(fill='x')

        code_frame = tk.Frame(body, bg=XP['panel_bg'])
        code_frame.pack(fill='x', pady=(0, 12))
        tk.Label(code_frame, text="Codigo (code):", bg=XP['panel_bg'], fg=XP['text'],
                 font=("Tahoma", 9, "bold"), width=14, anchor='w').pack(side='left')
        code_var = tk.StringVar()
        code_entry = tk.Entry(code_frame, textvariable=code_var, font=("Tahoma", 9),
                              bg=XP['field_bg'], relief='sunken', bd=2)
        code_entry.pack(side='left', fill='x', expand=True)
        code_entry.focus_set()

        btn_frame = tk.Frame(body, bg=XP['panel_bg'])
        btn_frame.pack()

        def on_generate():
            code = code_var.get().strip()
            win.destroy()
            self._run_odoo_bom(variant_name, bom_xlsx, code)

        def on_cancel():
            win.destroy()

        XPButton(btn_frame, text="Generar CSV", command=on_generate, width=120, height=28).pack(side='left', padx=6)
        XPButton(btn_frame, text="Cancelar", command=on_cancel, width=100, height=28).pack(side='left', padx=6)
        win.bind('<Return>', lambda e: on_generate())
        win.bind('<Escape>', lambda e: on_cancel())
        win.update_idletasks()
        x = self.winfo_x() + (self.winfo_width() - win.winfo_width()) // 2
        y = self.winfo_y() + (self.winfo_height() - win.winfo_height()) // 2
        win.geometry(f"+{x}+{y}")

    def _run_odoo_bom(self, variant_name, bom_xlsx, code):
        self._log(f"\n{'='*50}", 'dim')
        self._log("Generando BOM para Odoo...", 'info')
        self.status_var.set("Generando BOM Odoo...")
        try:
            csv_path = generate_odoo_bom_csv(bom_xlsx, code, lambda m: self._log(m, 'dim'))
            if csv_path:
                self._log(f"BOM Odoo: {os.path.basename(csv_path)}", 'ok')
                self.status_var.set(f"BOM Odoo generado: {os.path.basename(csv_path)}")
            else:
                self._log("BOM Odoo: sin componentes en el BOM", 'warn')
                self.status_var.set("BOM Odoo: sin datos")
        except Exception as e:
            self._log(f"Error generando BOM Odoo: {e}", 'err')
            self.status_var.set("BOM Odoo: error")


if __name__ == '__main__':
    app = KiBotGUI()
    app.mainloop()
